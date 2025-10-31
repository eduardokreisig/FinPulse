"""Excel workbook utilities."""

import logging
from copy import copy as _copy
from pathlib import Path
from typing import Any, Dict, Tuple

from openpyxl import load_workbook

from ..utils.path_utils import validate_path


def header_map(ws: Any) -> Dict[str, int]:
    """Create mapping of header names to column indices."""
    mapping = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip():
            mapping[v.strip()] = c
    return mapping


# Alias for backward compatibility
header_to_index = header_map


def find_insert_index(ws: Any, key: Tuple[str, str, Any], idx_bank: int, idx_account: int, idx_date: int) -> int:
    """Find the correct insertion index for sorted data."""
    lo, hi = 2, ws.max_row + 1
    while lo < hi:
        mid = (lo + hi) // 2
        mb = (ws.cell(row=mid, column=idx_bank).value or "")
        ma = (ws.cell(row=mid, column=idx_account).value or "")
        md = ws.cell(row=mid, column=idx_date).value
        if hasattr(md, "date"):
            md = md.date()
        if (mb, ma, md) <= key:
            lo = mid + 1
        else:
            hi = mid
    return lo


def clone_cell_style(src: Any, dst: Any) -> None:
    """Clone cell styling from source to destination."""
    dst.font = _copy(src.font)
    dst.border = _copy(src.border)
    dst.fill = _copy(src.fill)
    dst.number_format = src.number_format
    dst.protection = _copy(src.protection)
    dst.alignment = _copy(src.alignment)


def copy_row_styles(ws: Any, from_row: int, to_row: int) -> None:
    """Copy row styling from one row to another."""
    for c in range(1, ws.max_column + 1):
        clone_cell_style(ws.cell(row=from_row, column=c), ws.cell(row=to_row, column=c))


def should_skip_write(ws: Any, row_idx: int, col_idx: int) -> bool:
    """Check if we should skip writing to a cell (e.g., if neighbors have formulas)."""
    for neighbor in (row_idx - 1, row_idx + 1):
        if 2 <= neighbor <= ws.max_row:
            vv = ws.cell(row=neighbor, column=col_idx).value
            if isinstance(vv, str) and vv.startswith("="):
                return True
    return False


def load_workbook_safe(xlsx_path: Path):
    """Load workbook with path validation and error handling."""
    try:
        validated_xlsx = validate_path(xlsx_path)
        wb = load_workbook(validated_xlsx)
        return wb, validated_xlsx
    except (FileNotFoundError, PermissionError, ValueError) as e:
        logging.error(f"Failed to load workbook {xlsx_path}: {e}")
        raise


def save_workbook_safe(wb: Any, xlsx_path: Path) -> None:
    """Save workbook with error handling."""
    try:
        wb.save(xlsx_path)
        wb.close()
    except (PermissionError, OSError) as e:
        logging.error(f"Failed to save workbook {xlsx_path}: {e}")
        wb.close()
        raise