"""Excel sheet insertion operations."""

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl.formula.translate import Translator


from ..utils.date_utils import DATE_SEARCH_PATTERN, to_iso_dateish
from .workbook import (
    copy_row_styles, find_insert_index, header_map, header_to_index,
    load_workbook_safe, save_workbook_safe, should_skip_write
)


def is_safe_formula(formula: str) -> bool:
    """Check if formula is safe to use (no dangerous functions)."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return False
    
    # List of potentially dangerous Excel functions
    dangerous_functions = [
        'INDIRECT', 'OFFSET', 'INDEX', 'EXEC', 'CALL', 'REGISTER',
        'EVALUATE', 'HYPERLINK', 'WEBSERVICE', 'FILTERXML'
    ]
    
    formula_upper = formula.upper()
    return not any(func in formula_upper for func in dangerous_functions)


def sanitize_row_number(row_num: int, max_row: int = 1048576) -> int:
    """Sanitize row number to prevent injection."""
    if not isinstance(row_num, int) or row_num < 1 or row_num > max_row:
        raise ValueError(f"Invalid row number: {row_num}")
    return row_num


def copy_formulas_to_row(ws, src_row: int, dst_row: int, max_col: int = None) -> None:
    """Copy and translate formulas from source row to destination row."""
    # Validate inputs
    src_row = sanitize_row_number(src_row)
    dst_row = sanitize_row_number(dst_row)
    
    max_col = max_col or ws.max_column
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        
        if isinstance(src.value, str) and src.value.startswith("="):
            # Only process safe formulas
            try:
                if not is_safe_formula(src.value):
                    logging.warning(f"Skipping unsafe formula at {src.coordinate}: {src.value[:50]}...")
                    continue
            except Exception as e:
                logging.error(f"Formula safety check failed: {e}")
                continue
                
            try:
                # Only translate simple formulas
                if len(src.value) < 100:
                    dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
                else:
                    # For complex formulas, just copy as-is to avoid corruption
                    dst.value = src.value
            except (AttributeError, ValueError, KeyError, TypeError) as e:
                # If translation fails, copy the original formula
                dst.value = src.value
                logging.debug(f"Skipped formula translation for {src.coordinate}: {e}")


def safe_set_cell(ws, row: int, col: int, value: Any) -> None:
    """Safely set cell value if column exists and write is allowed."""
    if col and not should_skip_write(ws, row, col):
        # Sanitize inputs
        try:
            row = sanitize_row_number(row)
            if isinstance(value, str) and value.startswith("="):
                if not is_safe_formula(value):
                    logging.warning(f"Blocked unsafe formula: {value[:50]}...")
                    return
            ws.cell(row=row, column=col).value = value
        except (ValueError, TypeError) as e:
            logging.error(f"Failed to set cell value: {e}")
            return


def calculate_amount_from_withdrawals_deposits(w_amt: float, d_amt: float) -> float:
    """Calculate net amount from withdrawal and deposit amounts."""
    # With negative withdrawals format: just return whichever is non-zero
    w_amt = float(w_amt or 0.0)
    d_amt = float(d_amt or 0.0)
    return w_amt if w_amt != 0.0 else d_amt


def build_dedup_key(bank: str, account: str, date, description: str, amount: float) -> tuple:
    """Build standardized deduplication key with consistent date format."""
    date_str = to_iso_dateish(date)
    return (str(bank), str(account), date_str, norm_key(description), round(float(amount), 2))


def build_key_from_row_data(bank_label: str, account_label: str, row_data: dict) -> tuple:
    """Centralized function to build dedup key from normalized row data."""
    dkey = row_data["date"].date() if hasattr(row_data["date"], "date") else row_data["date"]
    return build_dedup_key(bank_label, account_label, dkey, row_data["description"], row_data["amount"])


def norm_key(s: Optional[str]) -> str:
    """Normalize key for comparison."""
    if s is None:
        return ""
    return " ".join(str(s).split()).lower()


def fix_shifted_formulas(ws, col_acc_period: int) -> None:
    """Fix formulas that were incorrectly shifted by Excel's automatic adjustment."""
    if not col_acc_period:
        return
    
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_acc_period)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            # Look for DATE formulas with cell references
            pattern = r'=DATE\(YEAR\(C(\d+)\),\s*MONTH\(C(\d+)\),\s*1\)'
            match = re.match(pattern, cell.value)
            if match:
                try:
                    ref_row1, ref_row2 = int(match.group(1)), int(match.group(2))
                    # Validate row numbers
                    sanitize_row_number(ref_row1)
                    sanitize_row_number(ref_row2)
                    sanitize_row_number(row_idx)
                    
                    # If both references point to the same row but not the current row, fix it
                    if ref_row1 == ref_row2 and ref_row1 != row_idx:
                        corrected_formula = f"=DATE(YEAR(C{row_idx}), MONTH(C{row_idx}), 1)"
                        if is_safe_formula(corrected_formula):
                            cell.value = corrected_formula
                except (ValueError, TypeError) as e:
                    logging.warning(f"Invalid row reference in formula: {e}")
                    continue


def insert_into_details(xlsx_path: Path, sheet_name: str, bank_label: str, 
                       account_label: str, rows: List[Dict[str, Any]], dry: bool = False,
                       cumulative_keys: dict = None, log_dir: Path = None) -> tuple[int, int, set]:
    """Insert rows into the Details sheet."""
    if not rows:
        return 0, 0, set()
    
    wb, validated_xlsx = load_workbook_safe(xlsx_path)
    
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"Sheet '{sheet_name}' not found")
    
    ws = wb[sheet_name]
    h = header_map(ws)

    col_bank = h.get("Bank")
    col_account = h.get("Account")
    col_date = h.get("Date")
    col_desc = h.get("Transaction Description")
    col_w = h.get("Withdrawals")
    col_d = h.get("Deposits")
    col_type = h.get("Transaction Type")
    col_acc_period = h.get("Accrual period")
    col_rev = h.get("Reviewed by Eduardo")
    col_notes = h.get("Notes")
    col_type_manual = h.get("Type")

    # Initialize debug log
    debug_log = []
    debug_log.append(f"=== DETAILS SHEET DEBUG: {bank_label} {account_label} ===")
    
    # Always build from existing Excel data, then merge with cumulative keys
    existing_keys = set()
    
    # Add cumulative keys if provided
    if cumulative_keys and 'details' in cumulative_keys and cumulative_keys['details']:
        existing_keys.update(cumulative_keys['details'])
        debug_log.append(f"Added {len(cumulative_keys['details'])} cumulative keys")
    
    # Always also check existing Excel data for current account
    for row_idx in range(2, ws.max_row + 1):
            b = (ws.cell(row=row_idx, column=col_bank).value or "")
            a = (ws.cell(row=row_idx, column=col_account).value or "")
            d = ws.cell(row=row_idx, column=col_date).value
            desc = ws.cell(row=row_idx, column=col_desc).value
            
            # Skip empty rows - check if Bank and Account have data (minimum required)
            if not b and not a:
                continue
            
            # Only include existing keys for the current bank/account being processed
            if str(b) != bank_label or str(a) != account_label:
                debug_log.append(f"Skipping row {row_idx}: Bank='{b}' vs '{bank_label}', Account='{a}' vs '{account_label}'")
                continue
            else:
                debug_log.append(f"Including row {row_idx}: Bank='{b}', Account='{a}'")
                
            if hasattr(d, "date"):
                d = d.date()
            # Get amount for deduplication key
            w_amt = ws.cell(row=row_idx, column=col_w).value or 0.0
            d_amt = ws.cell(row=row_idx, column=col_d).value or 0.0
            amt = calculate_amount_from_withdrawals_deposits(w_amt, d_amt)
            key = build_dedup_key(b, a, d, desc, amt)
            existing_keys.add(key)
            debug_log.append(f"Added existing key: {key}")

    rows_sorted = sorted(rows, key=lambda r: (bank_label, account_label, r["date"]))
    added = 0
    new_keys = set()
    
    debug_log.append(f"Existing keys count: {len(existing_keys)}")
    debug_log.append(f"Sample existing keys: {list(existing_keys)[:3]}")
    debug_log.append(f"Looking for Bank='{bank_label}', Account='{account_label}'")
    
    # Debug: Show what Bank/Account values actually exist
    found_accounts = set()
    for row_idx in range(2, min(ws.max_row + 1, 20)):  # Check first 20 rows
        b = ws.cell(row=row_idx, column=col_bank).value
        a = ws.cell(row=row_idx, column=col_account).value
        if b or a:
            found_accounts.add((repr(b), repr(a)))
    debug_log.append(f"Found Bank/Account pairs: {list(found_accounts)[:10]}")
    
    debug_log.append(f"Sample existing keys: {list(existing_keys)[:3]}")
    
    for row_data in rows_sorted:
        key = build_key_from_row_data(bank_label, account_label, row_data)
        debug_log.append(f"New key: {key}")
        if key in existing_keys:
            debug_log.append(f"  -> DUPLICATE (skipping)")
            continue
        debug_log.append(f"  -> NEW (adding)")

        if dry:
            added += 1
            continue

        dkey = row_data["date"].date() if hasattr(row_data["date"], "date") else row_data["date"]
        ins_at = find_insert_index(ws, (bank_label, account_label, dkey), h["Bank"], h["Account"], h["Date"])
        ws.insert_rows(ins_at, 1)
        src_row = ins_at - 1 if ins_at > 2 else ins_at + 1
        copy_row_styles(ws, src_row, ins_at)
        
        # Copy formulas from source row
        copy_formulas_to_row(ws, src_row, ins_at)

        safe_set_cell(ws, ins_at, col_bank, bank_label)
        safe_set_cell(ws, ins_at, col_account, account_label)
        safe_set_cell(ws, ins_at, col_date, dkey)
        safe_set_cell(ws, ins_at, col_desc, row_data["description"])

        amt = float(row_data["amount"] or 0.0)
        if amt < 0:
            safe_set_cell(ws, ins_at, col_w, amt)  # Keep negative value
            safe_set_cell(ws, ins_at, col_d, 0.0)
        else:
            safe_set_cell(ws, ins_at, col_w, 0.0)
            safe_set_cell(ws, ins_at, col_d, amt)

        if col_acc_period and hasattr(dkey, "replace"):
            safe_set_cell(ws, ins_at, col_acc_period, dkey.replace(day=1))
        if col_type:
            # Validate row number before creating formula
            safe_row = sanitize_row_number(ins_at)
            formula = f'=IF(AND(F{safe_row}<>0,G{safe_row}<>0),"Error",IF(F{safe_row}<0,"Withdrawal",IF(G{safe_row}>0,"Deposit","")))'  
            if is_safe_formula(formula):
                ws.cell(row=ins_at, column=col_type).value = formula
        safe_set_cell(ws, ins_at, col_rev, "No")
        safe_set_cell(ws, ins_at, col_notes, None)
        safe_set_cell(ws, ins_at, col_type_manual, None)

        added += 1
        new_keys.add(key)
        
        # Skip formula updates in existing rows to prevent corruption
        # Excel will automatically adjust most formulas when rows are inserted

    # Fix any formulas that were incorrectly shifted by Excel
    # Write debug log
    debug_log.append(f"Final: added={added}, existing_keys={len(existing_keys)}")
    try:
        if log_dir:
            debug_dir = log_dir / "Debug"
            debug_dir.mkdir(parents=True, exist_ok=True)
            debug_path = debug_dir / f"details_debug_{bank_label.replace(' ', '_')}_{account_label.replace(' ', '_')}.log"
            with open(debug_path, "w") as f:
                f.write("\n".join(debug_log))
            print(f"Debug log: {debug_path}")
    except Exception as e:
        print(f"Failed to write debug log: {e}")
    
    if not dry and added > 0:
        fix_shifted_formulas(ws, col_acc_period)
        save_workbook_safe(wb, validated_xlsx)
    elif not dry:
        save_workbook_safe(wb, validated_xlsx)
    return added, len(existing_keys), new_keys


def insert_into_account_sheet(xlsx_path: Path, sheet_name: str, bank_label: str, account_label: str,
                             rows: List[Dict[str, Any]], raw_map: Optional[Dict[str, str]], 
                             source_config: Optional[Dict[str, Any]] = None, dry: bool = False, 
                             start_date=None, end_date=None, cumulative_keys: dict = None, log_dir: Path = None) -> tuple[int, int, set]:
    """Insert rows into account-specific sheet."""
    if not rows:
        return 0, 0, set()
    
    wb, validated_xlsx = load_workbook_safe(xlsx_path)
    
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"Sheet '{sheet_name}' not found")
    
    ws = wb[sheet_name]
    h = header_to_index(ws)

    col_bank = h.get("Bank")
    col_account = h.get("Account")

    raw_cols_indices = {}
    for c in range(11, ws.max_column + 1):
        name = ws.cell(row=1, column=c).value
        if isinstance(name, str) and name.strip():
            raw_cols_indices[name.strip()] = c

    # Build existing dedupe keys using YAML config or fallback to hardcoded priority
    chosen = None
    
    # Try YAML-configured columns first
    if source_config:
        date_col = source_config.get("date_col")
        desc_col = source_config.get("description_col")
        
        if date_col and desc_col and date_col in raw_cols_indices and desc_col in raw_cols_indices:
            # Check for amount vs debit/credit
            if "Amount" in raw_cols_indices:
                chosen = [date_col, desc_col, "Amount"]
            elif source_config.get("debit_col") and source_config.get("credit_col"):
                debit_col = source_config.get("debit_col")
                credit_col = source_config.get("credit_col")
                if debit_col in raw_cols_indices and credit_col in raw_cols_indices:
                    chosen = [date_col, desc_col, debit_col, credit_col]
    
    # Fallback to hardcoded priority if YAML config didn't work
    if not chosen:
        prefer = [
            ["Date", "Description", "Amount"],
            ["Date", "Description", "Debit", "Credit"],
            ["Posting Date", "Description", "Amount"],
            ["Posting Date", "Description", "Debit", "Credit"],
            ["Post Date", "Description", "Amount"],
            ["Post Date", "Description", "Debit", "Credit"],
        ]
        for group in prefer:
            if all(n in raw_cols_indices for n in group):
                chosen = group
                break
    
    logging.info(f"Account sheet dedup columns for {bank_label} {account_label}: {chosen}")

    def get_existing_key_from_row(row_idx: int):
        # Use Bank/Account from main columns for consistency
        bank_val = ws.cell(row=row_idx, column=col_bank).value if col_bank else bank_label
        account_val = ws.cell(row=row_idx, column=col_account).value if col_account else account_label
        
        date_val = None
        desc_val = None
        amt_val = 0.0
        
        for n in chosen or []:
            cell_val = ws.cell(row=row_idx, column=raw_cols_indices[n]).value
            if DATE_SEARCH_PATTERN.search(n):
                date_val = cell_val
            elif n == "Description":
                desc_val = cell_val
            elif n == "Amount":
                amt_val = float(cell_val or 0)
            elif n in ["Debit", "Credit"]:
                if n == "Debit":
                    amt_val = float(cell_val or 0)
                elif n == "Credit" and amt_val == 0:
                    amt_val = float(cell_val or 0)
        
        if hasattr(date_val, "date"):
            date_val = date_val.date()
        
        return build_dedup_key(str(bank_val or bank_label), str(account_val or account_label), date_val, desc_val, amt_val)

    # Use cumulative keys if provided, otherwise build from existing data
    if cumulative_keys and 'accounts' in cumulative_keys and sheet_name in cumulative_keys['accounts']:
        existing = cumulative_keys['accounts'][sheet_name].copy()
        existing_in_range = len(existing)  # Approximate count
    else:
        existing = set()
        existing_in_range = 0
        if chosen:
            for row_index in range(2, ws.max_row + 1):
                key = get_existing_key_from_row(row_index)
                existing.add(key)
            
            # Count existing records in date range
            if start_date or end_date:
                date_col_name = None
                for n in chosen:
                    if DATE_SEARCH_PATTERN.search(n):
                        date_col_name = n
                        break
                
                if date_col_name:
                    for row_index in range(2, ws.max_row + 1):
                        date_val = ws.cell(row=row_index, column=raw_cols_indices[date_col_name]).value
                        try:
                            row_date = pd.to_datetime(date_val)
                            if start_date and row_date < pd.to_datetime(start_date):
                                continue
                            if end_date and row_date > pd.to_datetime(end_date):
                                continue
                            existing_in_range += 1
                        except (ValueError, TypeError):
                            pass
            else:
                existing_in_range = len(existing)
                
            logging.info(f"Found {len(existing)} existing records for dedup in {bank_label} {account_label}")
        else:
            logging.warning(f"No dedup columns found for {bank_label} {account_label}")

    # Find actual last row with data
    last_data_row = 1
    for r_idx in range(ws.max_row, 1, -1):
        bank_val = ws.cell(row=r_idx, column=col_bank).value if col_bank else None
        account_val = ws.cell(row=r_idx, column=col_account).value if col_account else None
        if bank_val or account_val:
            last_data_row = r_idx
            break
    
    # Find a row with formulas to copy from
    formula_source_row = last_data_row
    for r_idx in range(last_data_row, 1, -1):
        has_formula = False
        for c in range(1, min(11, ws.max_column + 1)):
            c_val = ws.cell(row=r_idx, column=c).value
            if isinstance(c_val, str) and c_val.startswith("="):
                has_formula = True
                break
        if has_formula:
            formula_source_row = r_idx
            break

    debug_log = []
    debug_log.append(f"=== ACCOUNT SHEET DEBUG: {bank_label} {account_label} ===")
    debug_log.append(f"Existing keys count: {len(existing)}")
    debug_log.append(f"Sample existing keys: {list(existing)[:3]}")
    
    added = 0
    new_keys = set()
    for row_data in rows:
        key = build_key_from_row_data(bank_label, account_label, row_data)
        debug_log.append(f"New key: {key}")
        if key in existing:
            debug_log.append(f"  -> DUPLICATE (skipping)")
            logging.debug(f"Skipping duplicate: {key}")
            continue
        debug_log.append(f"  -> NEW (adding)")

        if dry:
            added += 1
            continue

        # Insert after last data row
        ins_at = last_data_row + 1
        ws.insert_rows(ins_at, 1)
        
        # Copy styles and formulas from a row that has formulas
        copy_row_styles(ws, formula_source_row, ins_at)
        copy_formulas_to_row(ws, formula_source_row, ins_at, min(ws.max_column, 10))
        
        last_data_row = ins_at  # Update for next insertion

        safe_set_cell(ws, ins_at, col_bank, bank_label)
        safe_set_cell(ws, ins_at, col_account, account_label)

        for sheet_header, csv_col in (raw_map or {}).items():
            ci = raw_cols_indices.get(sheet_header)
            if not ci:
                continue
            cell_value = row_data.get("__raw__", {}).get(csv_col)
            if DATE_SEARCH_PATTERN.search(sheet_header):
                try:
                    cell_value = pd.to_datetime(cell_value).date()
                except (ValueError, TypeError):
                    pass
            safe_set_cell(ws, ins_at, ci, cell_value)

        added += 1
        new_keys.add(key)

    # Write debug log
    debug_log.append(f"Final: added={added}, existing_in_range={existing_in_range}")
    try:
        if log_dir:
            debug_dir = log_dir / "Debug"
            debug_dir.mkdir(parents=True, exist_ok=True)
            debug_path = debug_dir / f"account_debug_{bank_label.replace(' ', '_')}_{account_label.replace(' ', '_')}.log"
            with open(debug_path, "w") as f:
                f.write("\n".join(debug_log))
            print(f"Debug log: {debug_path}")
    except Exception as e:
        print(f"Failed to write debug log: {e}")
    
    if not dry:
        save_workbook_safe(wb, validated_xlsx)
    return added, existing_in_range, new_keys