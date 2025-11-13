"""
Handles ML inference workflow after data ingestion.

Responsibilities:
 - Load trained models and encoder
 - Perform inference for unlabeled rows
 - Write predictions into Category (K) and Subcategory (L) columns
 - Log summary statistics
"""

import pandas as pd
import yaml
import joblib
from pathlib import Path
from .preprocess import load_and_prepare_details


def run_ml_pipeline(cfg, xlsx_path: str):
    """Executes ML inference on timestamped workbook."""
    print("\n🔍 Starting FinPulse ML inference pipeline...")

    ml_cfg = cfg.get("ml", {})
    models_dir = Path(__file__).parent / "models"
    metadata_path = models_dir / "metadata.yaml"

    # Check for trained models
    if not metadata_path.exists():
        print("⚠️ No trained models found. Please run 'finpulse ml train' first.")
        return

    with open(metadata_path, "r") as f:
        meta = yaml.safe_load(f)
    version = meta["version"]

    # Check all required model files exist
    required_files = [
        models_dir / f"category_encoder_v{version}.joblib",
        models_dir / f"subcategory_encoder_v{version}.joblib",
        models_dir / f"category_v{version}.joblib",
        models_dir / f"subcategory_v{version}.joblib",
    ]
    missing = [f.name for f in required_files if not f.exists()]
    if missing:
        print(f"⚠️ Missing model files: {', '.join(missing)}. Re-train using 'finpulse ml train'.")
        return

    # Load models and encoders
    from .text_encoder import TextEncoder
    
    # Load text encoders properly
    category_encoder = TextEncoder(method=meta.get("global", {}).get("encoder", "tfidf"))
    category_encoder.load(str(models_dir / f"category_encoder_v{version}.joblib"))
    
    subcategory_encoder = TextEncoder(method=meta.get("global", {}).get("encoder", "tfidf"))
    subcategory_encoder.load(str(models_dir / f"subcategory_encoder_v{version}.joblib"))
    
    category_model = joblib.load(models_dir / f"category_v{version}.joblib")
    subcategory_model = joblib.load(models_dir / f"subcategory_v{version}.joblib")

    # Load workbook
    labeled_df, unlabeled_df = load_and_prepare_details(xlsx_path)
    df = pd.read_excel(xlsx_path, sheet_name="Details")

    # Filter unlabeled
    mask_unlabeled = df["Category"].isna() | df["Subcategory"].isna()
    df_unlabeled = df[mask_unlabeled].copy()

    if df_unlabeled.empty:
        print("✅ No unlabeled transactions found. Nothing to predict.")
        return

    # Get feature configurations
    category_features = ml_cfg.get("category_model", {}).get("features", ["Transaction Description", "Transaction Type"])
    subcategory_features = ml_cfg.get("subcategory_model", {}).get("features", ["Transaction Description", "Automated Trans. Category", "Transaction Type"])
    
    def build_text_features(df, feature_columns):
        """Build concatenated text features from specified columns."""
        if not feature_columns:
            return pd.Series([""] * len(df))
        
        text_series = df[feature_columns[0]].astype(str).fillna("")
        for col in feature_columns[1:]:
            if col in df.columns:
                text_series = text_series + " " + df[col].astype(str).fillna("")
        return text_series.str.lower()
    
    X_category_text = build_text_features(df_unlabeled, category_features)
    X_subcategory_text = build_text_features(df_unlabeled, subcategory_features)

    X_category = category_encoder.transform(X_category_text)
    X_subcategory = subcategory_encoder.transform(X_subcategory_text)

    # Predict categories and subcategories
    preds_category = category_model.predict(X_category)
    preds_subcategory = subcategory_model.predict(X_subcategory)

    df.loc[mask_unlabeled, "Category"] = preds_category
    df.loc[mask_unlabeled, "Subcategory"] = preds_subcategory

    # Save results preserving formatting
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb["Details"]
    
    # Find column indices
    col_category = col_subcategory = None
    for c in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=c).value
        if header == "Category":
            col_category = c
        elif header == "Subcategory":
            col_subcategory = c
    
    # Update only the prediction cells
    unlabeled_indices = df_unlabeled.index
    for i, (pred_category, pred_subcategory) in enumerate(zip(preds_category, preds_subcategory)):
        row_idx = unlabeled_indices[i] + 2  # +2 for 1-based indexing and header
        if col_category:
            ws.cell(row=row_idx, column=col_category).value = pred_category
        if col_subcategory:
            ws.cell(row=row_idx, column=col_subcategory).value = pred_subcategory
    
    wb.save(xlsx_path)
    wb.close()

    print(f"✅ ML predictions written to '{Path(xlsx_path).name}'.")
    print(f"   Category filled by ML: {len(preds_category)} rows")
    print(f"   Subcategory filled by ML: {len(preds_subcategory)} rows")
