"""Core data processing logic."""

import logging
from pathlib import Path
from typing import Tuple

import pandas as pd
from openpyxl import load_workbook

from ..data.csv_reader import load_inputs_by_file
from ..data.normalizer import normalize
from ..excel.sheet_inserter import insert_into_account_sheet, insert_into_details


def process_source(src_name: str, scfg: dict, xlsx: Path, details_sheet: str, args, cumulative_keys: dict = None) -> Tuple[int, int, int, int, int, dict]:
    """Process a single data source by processing each CSV file individually."""
    print(f"\n=== Source: {src_name} ===")
    csv_frames = load_inputs_by_file(scfg)
    
    if not csv_frames:
        # Still count existing records even if no new data to ingest
        account_sheet = scfg.get("account_sheet")
        if account_sheet and xlsx.exists():
            try:
                wb = load_workbook(xlsx, read_only=True)
                if account_sheet in wb.sheetnames:
                    ws = wb[account_sheet]
                    # Count non-empty rows (skip header)
                    existing_count = 0
                    for row_idx in range(2, ws.max_row + 1):
                        # Check if row has any data
                        has_data = False
                        for col_idx in range(1, min(ws.max_column + 1, 15)):  # Check first 15 columns
                            if ws.cell(row=row_idx, column=col_idx).value:
                                has_data = True
                                break
                        if has_data:
                            existing_count += 1
                    wb.close()
                    print(f"  -> no new data, but found {existing_count} existing records")
                    return 0, 0, 0, 0, existing_count, {}
                wb.close()
            except Exception as e:
                logging.warning(f"Failed to count existing records for {src_name}: {e}")
        return 0, 0, 0, 0, 0, {}
    
    print(f"  total CSV files: {len(csv_frames)}")
    print(f"  total rows across all files: {sum(len(df) for df in csv_frames)}")

    
    account_sheet = scfg.get("account_sheet")
    raw_map = scfg.get("raw_map")
    bank_label = scfg.get("bank_label", src_name)
    account_label = scfg.get("account_label", src_name)
    
    # Get raw_map once for all files
    if scfg.get("auto_raw_from_sheet", False):
        try:
            if not xlsx.exists():
                print(f"  Warning: Workbook {xlsx} does not exist, skipping auto_raw_from_sheet")
                raw_map = {}
            else:
                wb = load_workbook(xlsx, read_only=True, data_only=False)
                if account_sheet not in wb.sheetnames:
                    wb.close()
                    print(f"  Warning: Account sheet '{account_sheet}' not found, skipping auto_raw_from_sheet")
                    raw_map = {}
                else:
                    ws = wb[account_sheet]
                    raw_map = {}
                    for c in range(11, ws.max_column + 1):
                        name = ws.cell(row=1, column=c).value
                        if isinstance(name, str) and name.strip():
                            raw_map[name.strip()] = name.strip()
                    wb.close()
        except (FileNotFoundError, PermissionError, KeyError) as e:
            logging.warning(f"Failed to load workbook for auto_raw_from_sheet: {e}")
            raw_map = {}
    
    print(f"  raw_map keys (K+ headers): {list(raw_map.keys()) if raw_map else 'None'}")
    
    # Process each CSV file individually
    total_acct_added = 0
    total_det_added = 0
    total_nat_count = 0
    total_deduped_count = 0
    total_existing = 0
    
    for file_idx, df_in in enumerate(csv_frames, 1):
        source_file = df_in["__source_file"].iloc[0] if not df_in.empty else "unknown"
        print(f"\n  Processing file {file_idx}/{len(csv_frames)}: {Path(source_file).name}")
        print(f"    rows in file: {len(df_in)}")
        print(f"    CSV headers: {list(df_in.columns)}")
        
        norm_df = normalize(df_in, scfg)
        print(f"    normalized rows: {len(norm_df)}")
        
        try:
            dates = pd.to_datetime(norm_df["date"], errors="coerce")
            nat_count = dates.isna().sum()
            print(f"    date dtype: {dates.dtype}, NaT count: {nat_count} of {len(dates)}")
            if dates.notna().any():
                print(f"    date min/max: {dates.min()} .. {dates.max()}")
            norm_df["date"] = dates
        except (ValueError, TypeError) as e:
            logging.warning(f"Failed to process dates for file {source_file}: {e}")
            print(f"    Warning: Date processing failed, skipping file")
            continue
        
        if args.start:
            start = pd.to_datetime(args.start)
            norm_df = norm_df[norm_df["date"] >= start]
        if args.end:
            end = pd.to_datetime(args.end)
            norm_df = norm_df[norm_df["date"] <= end]
        
        print(f"    rows after date filter: {len(norm_df)}")
        if norm_df.empty:
            continue
        
        rows = []
        for idx, nrow in norm_df.iterrows():
            raw_payload = {}
            if raw_map:
                for sheet_header, csv_col in raw_map.items():
                    raw_payload[csv_col] = df_in.loc[idx, csv_col] if csv_col in df_in.columns else None
            rows.append({
                "date": nrow["date"],
                "amount": float(nrow["amount"]),
                "description": nrow["description"],
                "__raw__": raw_payload
            })
        
        # Process this file's data
        acct_added, acct_existing, new_acct_keys = insert_into_account_sheet(
            xlsx, account_sheet, bank_label, account_label, rows, raw_map=raw_map, 
            source_config=scfg, dry=args.dry_run, start_date=args.start, end_date=args.end,
            cumulative_keys=cumulative_keys, log_dir=getattr(args, 'log_dir_path', None)
        )
        det_added, det_existing, new_det_keys = insert_into_details(
            xlsx, details_sheet, bank_label, account_label, rows, dry=args.dry_run,
            cumulative_keys=cumulative_keys, log_dir=getattr(args, 'log_dir_path', None)
        )
        
        print(f"    -> file added: account={acct_added}, details={det_added}")
        
        # Update cumulative keys immediately after each file
        if cumulative_keys:
            if new_det_keys:
                cumulative_keys['details'].update(new_det_keys)
            if new_acct_keys:
                if 'accounts' not in cumulative_keys:
                    cumulative_keys['accounts'] = {}
                if account_sheet not in cumulative_keys['accounts']:
                    cumulative_keys['accounts'][account_sheet] = set()
                cumulative_keys['accounts'][account_sheet].update(new_acct_keys)
        
        # Accumulate totals
        total_acct_added += acct_added
        total_det_added += det_added
        total_nat_count += nat_count
        total_deduped_count += len(rows) - det_added
        if file_idx == 1:  # Only count existing records once
            total_existing = acct_existing
    
    print(f"  TOTAL existing records: {total_existing}")
    print(f"  TOTAL -> per-account added: {total_acct_added}, details added: {total_det_added}")
    
    # Return empty new_keys since we've already updated cumulative_keys
    return total_acct_added, total_det_added, total_nat_count, total_deduped_count, total_existing, {}