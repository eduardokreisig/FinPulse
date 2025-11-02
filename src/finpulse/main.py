"""Main entry point for FinPulse financial data ingestion."""

import argparse
import logging
import sys
import traceback
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from .config.loader import get_log_directory, load_config
from .data.csv_reader import load_inputs_for_source
from .data.normalizer import normalize
from .excel.sheet_inserter import insert_into_account_sheet, insert_into_details
from .utils.logging_utils import Tee, setup_logging, utc_log_name
from .utils.path_utils import get_timestamp


def get_user_input(prompt: str, default: str = None) -> str:
    """Get user input with optional default value."""
    if default:
        response = input(f"{prompt} [{default}]: ").strip()
        return response if response else default
    return input(f"{prompt}: ").strip()


def get_yes_no(prompt: str, default: bool = False) -> bool:
    """Get yes/no input from user."""
    default_str = "Y/n" if default else "y/N"
    while True:
        response = input(f"{prompt} [{default_str}]: ").strip().lower()
        if not response:
            return default
        if response in ['y', 'yes']:
            return True
        if response in ['n', 'no']:
            return False
        print("Please enter 'y' or 'n'")


def setup_log_file(log_dir: Path, is_dry_run: bool):
    """Setup log file if log directory is specified."""
    if not log_dir:
        return None, None, None

    try:
        log_dir.mkdir(parents=True, exist_ok=True)
        logfile = log_dir / utc_log_name(is_dry_run)
        log_fp = open(logfile, "w", encoding="utf-8")
        orig_stdout = sys.stdout
        tee = Tee(orig_stdout, log_fp)
        sys.stdout = tee
        print(f"[logging to] {logfile}")
        return log_fp, orig_stdout, tee
    except (OSError, PermissionError) as e:
        logging.error(f"Failed to setup logging to {log_dir}: {e}")
        raise


def get_interactive_config():
    """Get configuration interactively from user."""
    print("\n=== FinPulse Interactive Setup ===")

    # Get config file
    config_file = get_user_input("Config file path", "config/config.yaml")
    config_path = Path(config_file)
    if not config_path.exists():
        print(f"Warning: Config file {config_path} does not exist")
        if not get_yes_no("Continue anyway?", False):
            sys.exit(1)

    # Get workspace folder
    workspace = get_user_input("Finance workspace folder", "..")
    workspace_path = Path(workspace).resolve()

    # Get workbook name
    workbook_name = get_user_input("Finance workbook filename", "FinanceWorkbook 2025.xlsx")

    # Get inputs folder
    default_inputs = str(workspace_path / "Inputs")
    inputs_folder = get_user_input("Inputs folder", default_inputs)

    # Get logs folder
    default_logs = str(workspace_path / "FinPulseImportLogs")
    logs_folder = get_user_input("Logs folder", default_logs)

    # Get date range with current year defaults
    current_year = datetime.now().year
    default_start = f"{current_year}-01-01"
    default_end = f"{current_year}-12-31"

    start_date = get_user_input("Start date (YYYY-MM-DD)", default_start)
    end_date = get_user_input("End date (YYYY-MM-DD)", default_end)

    return {
        'config': config_file,
        'workspace': workspace,
        'workbook': workbook_name,
        'inputs': inputs_folder,
        'logs': logs_folder,
        'start': start_date if start_date else None,
        'end': end_date if end_date else None
    }


def process_source(src_name: str, scfg: dict, xlsx: Path, details_sheet: str, args) -> tuple[int, int]:
    """Process a single data source."""
    print(f"\n=== Source: {src_name} ===")
    df_in = load_inputs_for_source(scfg)
    print(f"  df_in rows: {len(df_in)}")

    if df_in.empty:
        return 0, 0

    # show headers early for debugging
    print(f"  CSV headers (after parse): {list(df_in.columns)}")

    account_sheet = scfg.get("account_sheet")
    raw_map = scfg.get("raw_map")

    if scfg.get("auto_raw_from_sheet", False):
        try:
            if not xlsx.exists():
                print(f"  Warning: Workbook {xlsx} does not exist, skipping auto_raw_from_sheet")
                raw_map = {}
            else:
                wb = load_workbook(xlsx, read_only=True, data_only=False)
                if account_sheet not in wb.sheetnames:
                    wb.close()
                    print(f"  Warning: Account sheet '{account_sheet}' not found, skipping auto_raw_from_sheet")
                    raw_map = {}
                else:
                    ws = wb[account_sheet]
                    raw_map = {}
                    for c in range(11, ws.max_column + 1):
                        name = ws.cell(row=1, column=c).value
                        if isinstance(name, str) and name.strip():
                            raw_map[name.strip()] = name.strip()
                    wb.close()
        except (FileNotFoundError, PermissionError, KeyError) as e:
            logging.warning(f"Failed to load workbook for auto_raw_from_sheet: {e}")
            raw_map = {}

    print(f"  raw_map keys (K+ headers): {list(raw_map.keys()) if raw_map else 'None'}")

    norm_df = normalize(df_in, scfg)
    print(f"  normalized rows: {len(norm_df)}")

    try:
        dates = pd.to_datetime(norm_df["date"], errors="coerce")
        print(f"  date dtype: {dates.dtype}, NaT count: {dates.isna().sum()} of {len(dates)}")
        if dates.notna().any():
            print(f"  date min/max: {dates.min()} .. {dates.max()}")
        norm_df["date"] = dates
    except (ValueError, TypeError) as e:
        logging.warning(f"Failed to process dates for source {src_name}: {e}")
        print(f"  Warning: Date processing failed, skipping source")
        return 0, 0

    if args.start:
        start = pd.to_datetime(args.start)
        norm_df = norm_df[norm_df["date"] >= start]
    if args.end:
        end = pd.to_datetime(args.end)
        norm_df = norm_df[norm_df["date"] <= end]

    print(f"  rows after date filter: {len(norm_df)}")
    if norm_df.empty:
        return 0, 0

    rows = []
    for idx, nrow in norm_df.iterrows():
        raw_payload = {}
        if raw_map:
            for sheet_header, csv_col in raw_map.items():
                raw_payload[csv_col] = df_in.loc[idx, csv_col] if csv_col in df_in.columns else None
        rows.append({
            "date": nrow["date"],
            "amount": float(nrow["amount"]),
            "description": nrow["description"],
            "__raw__": raw_payload
        })

    bank_label = scfg.get("bank_label", src_name)
    account_label = scfg.get("account_label", src_name)

    acct_added = insert_into_account_sheet(
        xlsx, account_sheet, bank_label, account_label, rows, raw_map=raw_map, 
        source_config=scfg, dry=args.dry_run
    )
    det_added = insert_into_details(xlsx, details_sheet, bank_label, account_label, rows, dry=args.dry_run)
    print(f"  -> per-account added: {acct_added}, details added: {det_added}")

    return acct_added, det_added


def main() -> None:
    """Main entry point."""
    setup_logging()

    ap = argparse.ArgumentParser(
        description="FinPulse v5.3.17 (robust CSV defaults; no YAML change needed)"
    )
    ap.add_argument("--config", default=None)
    ap.add_argument("--start", help="YYYY-MM-DD", default=None)
    ap.add_argument("--end", help="YYYY-MM-DD", default=None)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--log-dir", default=None)
    ap.add_argument("--workspace", default=None)
    ap.add_argument("--workbook", default=None)
    ap.add_argument("--inputs", default=None)
    args = ap.parse_args()

    # Interactive mode if no config provided
    interactive_config = None
    if not args.config:
        interactive_config = get_interactive_config()
        args.config = interactive_config['config']
        args.start = args.start or interactive_config['start']
        args.end = args.end or interactive_config['end']
        args.log_dir = args.log_dir or interactive_config['logs']

    try:
        cfg = load_config(args.config)
    except FileNotFoundError:
        print(f"Error: Config file '{args.config}' not found")
        sys.exit(1)
    except Exception as e:
        print(f"Error loading config: {e}")
        sys.exit(1)

    # Override config with interactive inputs if provided
    if interactive_config:
        workspace_path = Path(interactive_config['workspace']).resolve()
        cfg['target_workbook'] = str(workspace_path / interactive_config['workbook'])
        cfg['log_dir'] = interactive_config['logs']

        # Update source file paths to use new inputs folder
        inputs_path = Path(interactive_config['inputs'])
        for src_name, src_cfg in cfg.get('sources', {}).items():
            if 'files' in src_cfg:
                # Replace the base path in file globs
                updated_files = []
                for file_pattern in src_cfg['files']:
                    # Extract the relative part after 'Inputs'
                    pattern_path = Path(file_pattern)
                    if 'Inputs' in pattern_path.parts:
                        inputs_idx = pattern_path.parts.index('Inputs')
                        relative_part = Path(*pattern_path.parts[inputs_idx+1:])
                        new_pattern = inputs_path / relative_part
                        updated_files.append(str(new_pattern))
                    else:
                        updated_files.append(file_pattern)
                src_cfg['files'] = updated_files

    log_dir = get_log_directory(cfg, args.log_dir)

    orig_stdout = sys.stdout
    log_fp = None
    tee = None

    try:
        log_fp, orig_stdout, tee = setup_log_file(log_dir, args.dry_run)

        xlsx = Path(cfg['target_workbook']).expanduser().resolve()
        details_sheet = cfg.get("details_sheet", "Details")

        print(f"Workbook: {xlsx}")
        if not xlsx.exists():
            print(f"  exists: False  size: 0 bytes")
            print(f"  modified: n/a")
            print(f"Warning: Target workbook '{xlsx}' does not exist")
            if not args.dry_run and not get_yes_no("Continue anyway?", False):
                sys.exit(1)
        else:
            print(f"  exists: True  size: {xlsx.stat().st_size} bytes")
            print(f"  modified: {get_timestamp(xlsx)}")

        total_details = 0
        total_accounts = 0
        sources_processed = 0
        sources_with_data = 0
        source_results = []

        for src_name, scfg in cfg["sources"].items():
            sources_processed += 1
            acct_added, det_added = process_source(src_name, scfg, xlsx, details_sheet, args)
            total_accounts += acct_added
            total_details += det_added
            source_results.append((src_name, acct_added, det_added))
            if acct_added > 0 or det_added > 0:
                sources_with_data += 1

        print(f"\nSummary: {sources_processed} sources processed, {sources_with_data} had data")
        print(f"per-account added={total_accounts}, details added={total_details}")
        if xlsx.exists():
            print(f"Modified (after): {get_timestamp(xlsx)}")
        
        # Check for deduplication discrepancies
        if total_accounts != total_details:
            print(f"\nWARNING: Deduplication mismatch - per-account: {total_accounts}, details: {total_details}")
            discrepancies = [(name, acct, det) for name, acct, det in source_results if acct != det]
            if discrepancies:
                print("Accounts with discrepancies:")
                for name, acct_added, det_added in discrepancies:
                    print(f"  {name}: per-account={acct_added}, details={det_added}")

        if args.dry_run:
            print("(dry-run) No changes written.")
            if total_accounts > 0 or total_details > 0:
                if get_yes_no("\nProceed with real import?", False):
                    # Re-run without dry-run
                    args.dry_run = False
                    print("\n=== REAL IMPORT ===")
                    total_details = 0
                    total_accounts = 0
                    final_results = []
                    for src_name, scfg in cfg["sources"].items():
                        acct_added, det_added = process_source(src_name, scfg, xlsx, details_sheet, args)
                        total_accounts += acct_added
                        total_details += det_added
                        final_results.append((src_name, acct_added, det_added))
                    print(f"\nFinal: per-account added={total_accounts}, details added={total_details}")
                    
                    # Check for deduplication discrepancies in final run
                    if total_accounts != total_details:
                        print(f"\nWARNING: Deduplication mismatch - per-account: {total_accounts}, details: {total_details}")
                        final_discrepancies = [(name, acct, det) for name, acct, det in final_results if acct != det]
                        if final_discrepancies:
                            print("Accounts with discrepancies:")
                            for name, acct_added, det_added in final_discrepancies:
                                print(f"  {name}: per-account={acct_added}, details={det_added}")
                    print("Done.")
        else:
            print("Done.")

    except KeyboardInterrupt:
        print("\nOperation cancelled by user")
        return
    except Exception as e:
        print("ERROR:", e)
        print(traceback.format_exc())
        return
    finally:
        # Restore stdout first
        if tee:
            tee.close()
        sys.stdout = orig_stdout

        # Then close the log file
        if log_fp and not log_fp.closed:
            try:
                log_fp.close()
            except (OSError, IOError):
                pass  # Ignore errors during cleanup


if __name__ == "__main__":
    main()